'''
This module handles the methods used for data processing.

Features:
- Decode, upload, load and treat VO2 and lactate contents.
- Modelize data.
- Compute the thresholds, the results and the plateau analysis.

'''

### Imports ###

import pandas as pd
from io import StringIO
import logging
import copy
from scipy.signal import find_peaks
from scipy import stats
from datetime import time, timedelta
import xml.etree.ElementTree as ET
import pwlf
import base64
import io
import json
from openpyxl import load_workbook
import numpy as np
from utils.database_updating import *
from sklearn.preprocessing import PolynomialFeatures
from sklearn.linear_model import LinearRegression
import pwlf
import numpy as np
from utils.models import *
import traceback
import math
from unidecode import unidecode
from scipy.integrate import quad
import plotly.graph_objects as go
from plotly.subplots import make_subplots

### Methods ###

def decode_contents(contents):
    """
    Decode a Base64-encoded file from a content string.

    This function takes a string containing metadata (content type) and 
    Base64-encoded data, separated by a comma. It returns an `io.BytesIO` 
    object representing the decoded data as an in-memory file.

    Parameters
    ----------
    contents : str
        A string containing metadata and Base64-encoded data in the following format:
        "data:<content_type>;base64,<encoded_data>"
        Example: "data:application/pdf;base64,SGVsbG8sIFdvcmxkIQ=="

    Returns
    -------
    io.BytesIO
        An in-memory file-like object containing the decoded data.
    """

    try:
        # Ensure the input string has the expected format
        if not contents or ',' not in contents:
            raise ValueError("Le contenu fourni n'est pas correctement formé.")

        # Split into content type and encoded data
        content_type, content_string = contents.split(',')

        # Decode the Base64 data
        decoded = base64.b64decode(content_string)
        file = io.BytesIO(decoded)

        return file

    except Exception as e:
        logging.error(f"Error in decode_contents: {e}\n{traceback.format_exc()}")
        raise
    
def upload_vo2(contents) :
    """
    Processes and validates a VO2 file encoded in Base64, extracts relevant data, 
    and converts it into a pandas DataFrame.

    Parameters
    ----------
    contents : str
        Base64-encoded content of the VO2 file.

    Returns
    -------
    tuple
        - pandas.DataFrame: Contains processed data if successful, otherwise `None`.
        - str: An empty error message if successful, or a descriptive error message if an issue occurs.
    """

    # Namespace definition for XML parsing
    NAMESPACE = '{urn:schemas-microsoft-com:office:spreadsheet}'

    try:

        # Decode the Base64 content and parse the XML
        tree = ET.parse(decode_contents(contents))
        root = tree.getroot()

        # Locate the worksheet element
        worksheet = root.find(f'{NAMESPACE}Worksheet')
        if worksheet is None:
            raise ValueError("Worksheet element not found in XML.")
        
        # Get the worksheet name
        worksheet_name = worksheet.attrib.get(f'{NAMESPACE}Name')
        if worksheet_name is None:
            raise ValueError("Worksheet name attribute not found.")
        
        # Locate the table element
        table = worksheet.find(f'{NAMESPACE}Table')
        if table is None:
            raise ValueError("Table element not found in worksheet.")
        
        # Extract rows from the table
        rows = table.findall(f'{NAMESPACE}Row')
        if not rows:
            raise ValueError("No rows found in the table.")
        
        # Extract data from each row
        data = []
        for row in rows:
            cells = row.findall(f'{NAMESPACE}Cell')
            row_data = [cell.findtext(f'{NAMESPACE}Data', default='') for cell in cells]
            data.append(row_data)
        
        # Ensure data is not empty
        if not data:
            raise ValueError("No data")

        # Convert the extracted data into a DataFrame
        df = pd.DataFrame(data)

        # Define configurations for supported worksheet types
        configurations = {
            "Metasoft": {
                "signal": "Temps de mesure",
                "columns_to_drop": [12, 13, 14],
                "columns_name": ["t", "FR", "VE", "VO2", "VCO2", "VO2/kg", "PetO2", "PetCO2", "RER", "VE/VO2", "VE/VCO2", "FC"]
            },
            "MetasoftStudio": {
                "signal": "t",
                "columns_to_drop": [1, 2, 5, 7, 13, 14, 16],
                "columns_name": ["t", "VO2", "VO2/kg", "FC", "VE/VO2", "VE/VCO2", "RER", "VE", "FR", "VCO2"]
            }
        }

        # Retrieve configuration based on worksheet name
        if worksheet_name not in configurations:
            raise ValueError(f"Unknown worksheet name: {worksheet_name}")
        
        config = configurations[worksheet_name]
        signal = config["signal"]
        columns_to_drop = config["columns_to_drop"]
        columns_name = config["columns_name"]

        # Identify the starting index of test data
        try:
            index_limit = df.index[df.iloc[:, 0] == signal].tolist()[0] + 2
            df = df.loc[index_limit:]
        except IndexError:
            raise ValueError(f"Signal '{signal}' not found in the data.")
        
        # Drop unnecessary columns and reindex the DataFrame
        df = df.drop(columns=columns_to_drop).dropna().reset_index(drop=True)

        # Rename columns
        df.columns = columns_name

        # Convert numeric columns to float
        df.iloc[:, 1:] = df.iloc[:, 1:].astype(float)

        return df, ""

    except Exception as e:
        logging.error(f"Error in upload_vo2: {e}\n{traceback.format_exc()}")
        raise
    
def upload_lactate(contents) :
    """
    Processes and validates a lactate file encoded in Base64, extracts relevant data, 
    and converts it into a pandas DataFrame.

    Parameters
    ----------
    contents : str
        Base64-encoded content of the lactate file.

    Returns
    -------
    tuple
        - pandas.DataFrame: Contains processed data if successful, otherwise `None`.
        - str: An empty error message if successful, or a descriptive error message if an issue occurs.
        - dict: Athlete profile data extracted from the file, if applicable.
    """
    try:
        # Decode and load the Excel file
        try:
            workbook = load_workbook(decode_contents(contents), data_only=True).active
        except Exception as e:
            raise ValueError("Echec du téléchargement du fichier Excel.") from e
        
        # Extract data into a list of rows
        data = [row for row in workbook.iter_rows(values_only=True)]

        # Convert the data into a DataFrame, replacing None with NaN
        df_lactate = pd.DataFrame(data).replace(to_replace=[None], value=np.nan)

        # Extract athlete profile information
        data_athlete = load_athlete_profile(df_lactate)

        return df_lactate, "", data_athlete
    
    except Exception as e:
        logging.error(f"Error in upload_lactate: {e}\n{traceback.format_exc()}")
        raise
    
def treat_lactate(data_lactate):
    """
    Process lactate data from a JSON file.

    This function loads, cleans, and prepares lactate data encoded in JSON format. 
    It ensures the data contains the required columns, removes unnecessary values, 
    and interpolates missing values in numeric columns.

    Parameters
    ----------
    data_lactate : str
        JSON string representing lactate data.

    Returns
    -------
    tuple
        - pd.DataFrame: The processed DataFrame if the data is valid.
        - str: An empty string if successful, or a descriptive error message.
    """
    try:
        # Step 1: Load data from JSON
        try:
            df_lactate = pd.read_json(StringIO(data_lactate))
        except (ValueError, json.JSONDecodeError) as e:
            return None, f"Erreur lors du chargement des données JSON : {e}"
        
        # Step 2: Prepare the data
        # Validate minimum dimensions of the DataFrame
        if df_lactate.shape[0] < 10 or df_lactate.shape[1] < 12:
            return None, "Les données sont insuffisantes ou mal formatées."
        
        # Extract and clean columns
        df_lactate = df_lactate.iloc[9:, :12].reset_index(drop = True)
        df_lactate.columns = df_lactate.iloc[0]
        df_lactate = df_lactate.drop(df_lactate.index[0]).reset_index(drop = True)
        df_lactate.columns = [name.lower() for name in df_lactate.columns]

        # Step 3: Validate expected columns
        required_columns = {'lactate', 'fc'}
        if not required_columns.issubset(df_lactate.columns):
            return None, f"Colonnes manquantes : {required_columns - set(df_lactate.columns)}"

        # Step 4: Clean and interpolate data
        # Remove trailing null values at the end of the test
        df_lactate = trim_trailing_nulls(df_lactate, 'lactate')

        # Interpolate numeric columns (ensure they exist)
        numeric_columns = df_lactate.columns[5:8]
        if len(numeric_columns) > 0:
            df_lactate[numeric_columns] = df_lactate[numeric_columns].astype(float).interpolate()

        # Check for null values in the 'fc' column
        if df_lactate['fc'].isnull().all():
            raise ValueError("Toutes les valeurs de 'FC' sont nulles. Veuillez vérifier la source des données.")

        return df_lactate
    
    except Exception as e:
        logging.error(f"Error in treat_lactate: {e}\n{traceback.format_exc()}")
        raise

def trim_trailing_nulls(df, column_name):
    """
    Removes trailing null or empty values from a specific column in a DataFrame.

    This function iterates through the specified column from the end, identifying the last non-null 
    and non-empty value. All rows beyond this value are removed, effectively trimming trailing invalid data.

    Parameters
    ----------
    df : pandas.DataFrame
        The DataFrame to be processed.
    column_name : str
        The name of the column from which trailing null or empty values should be removed.

    Returns
    -------
    pandas.DataFrame
        A DataFrame without trailing null or empty values in the specified column.
    """
    try:
        # Identify the last valid index in the specified column
        last_index = len(df[column_name]) - 1
        
        # Traverse the column backwards to find the last non-null and non-empty value
        while last_index >= 0:
            cell_value = df[column_name].iloc[last_index]
            if pd.notna(cell_value) and cell_value != "":
                break
            last_index -= 1

        # Raise an error if no valid data exists in the column
        if last_index < 0:
            raise ValueError(f"No valid data found in '{column_name}' column.")

        return df.iloc[:last_index + 1].reset_index(drop=True)
    
    except Exception as e:
        logging.error(f"Error in trim_trailing_nulls: {e}\n{traceback.format_exc()}")
        raise

def format_columns(df, numeric_exclude=None):
    """
    Formats DataFrame columns for use in a Dash DataTable, specifying numeric formats.

    This function generates a list of dictionaries that define the formatting of columns
    in a Dash DataTable. Numeric columns are formatted with a precision of two decimal places,
    unless they are specified in the `numeric_exclude` list.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame whose columns need to be formatted.
    numeric_exclude : list, optional
        A list of column names to exclude from numeric formatting. Defaults to an empty list.

    Returns
    -------
    list of dict
        A list of dictionaries representing formatted column definitions for Dash DataTable.
    """
    try:
        # Default to an empty list if no exclusions are provided
        numeric_exclude = numeric_exclude or []

        return [
                    {
                        "name": col,
                        "id": col,
                        "type": "numeric",
                        "format": {"specifier": ".2f"},
                    } if col not in numeric_exclude else {"name": col, "id": col}
                    for col in df.columns
                ]
    
    except Exception as e:
        logging.error(f"Error in format_columns: {e}\n{traceback.format_exc()}")
        raise

def load_dataframe_vo2(data_json):
    """
    Loads and processes VO2 data from a JSON string into a DataFrame.

    This function parses a JSON string containing VO2 data, converts it into a Pandas DataFrame,
    and ensures that the time column (`t`) is properly formatted as datetime objects.

    Parameters
    ----------
    data_json : str
        A JSON string representing the VO2 data.

    Returns
    -------
    pd.DataFrame
        A DataFrame containing the processed VO2 data.
    """
    try:

        # Convert JSON string to DataFrame
        df = pd.read_json(StringIO(data_json), orient='records')

        # Convert the time column to datetime format
        df['t'] = pd.to_datetime(df['t'], format="mixed")

        return df
    
    except Exception as e:
        logging.error(f"Error in load_dataframe_vo2: {e}")
        raise
    
def load_dataframe_lactate(data_json):
    """
    Loads lactate data from a JSON string into a DataFrame.

    This function parses a JSON string containing lactate data and converts it into a Pandas DataFrame.

    Parameters
    ----------
    data_json : str
        A JSON string representing the lactate data.

    Returns
    -------
    pd.DataFrame
        A DataFrame containing the lactate data.
    """
    try:
        # Convert JSON string to DataFrame
        df = pd.read_json(StringIO(data_json), orient='records')
        return df
    
    except Exception as e:
        logging.error(f"Errorin load_dataframe_lactate: {e}\n{traceback.format_exc()}")
        raise
    
def load_dataframe_from_id(test_id, lactate=False):
    """
    Loads a DataFrame from a test ID, based on the test type.

    This function retrieves a test object from the database using its ID, and then loads the 
    appropriate DataFrame (VO2 or lactate) based on the provided parameters.

    Parameters
    ----------
    test_id : int
        The unique identifier of the test in the database.
    lactate : bool, optional
        If `True`, loads the lactate DataFrame; otherwise, loads the VO2 DataFrame. Default is `False`.

    Returns
    -------
    pd.DataFrame
        A DataFrame containing the data for the specified test.
    """
    try:
        # Retrieve the database session
        session = next(get_db())
        test = get_by_id(session, Test, test_id)

        # Load the appropriate DataFrame based on the test type
        if not lactate:
            df = load_dataframe_vo2(test.computed_dataframe)
        else:
            df = load_dataframe_lactate(test.source_excel)

        return df
    
    except Exception as e:
        logging.error(f"Errorin load_dataframe_from_id: {e}\n{traceback.format_exc()}")
        raise

def divise_df(df):
    """
    Splits a DataFrame into two parts based on the median value of the 't' column.

    The function calculates the median of the 't' column in the DataFrame and splits the data
    into two parts: one containing values less than or equal to the median, and the other containing
    values greater than the median.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame to be divided. It must include a column named 't'.

    Returns
    -------
    tuple
        - First part of the DataFrame (`pd.Series`) containing values <= median.
        - Second part of the DataFrame (`pd.Series`) containing values > median.
    """
    try:
        # Calculate the median value of the 't' column
        med_df = df['t'].median()

        # Divide the DataFrame into two parts based on the median
        first_part_df = df['t'][df['t'] <= med_df]
        second_part_df = df['t'][df['t'] > med_df].reset_index(drop=True)

        return first_part_df, second_part_df

    except Exception as e:
        logging.error(f"Error in divise_df: {e}\n{traceback.format_exc()}")
        raise

def remove_clicked_point(df, clicked_time):
    """
    Removes a row from the DataFrame based on a specified time value.

    This function filters out the row where the value in the 't' column matches the provided `clicked_time`.
    The DataFrame is then reset to maintain sequential indexing.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame from which a row will be removed. It must include a column named 't'.
    clicked_time : object
        The time value to be removed from the 't' column. The type should match the 't' column's datatype.

    Returns
    -------
    pd.DataFrame
        A DataFrame with the specified row removed and the index reset.
    """
    try:
        # Filter out the row with the specified 'clicked_time'
        df = df[df["t"] != clicked_time].reset_index(drop=True)
        return df
    
    except Exception as e:
        logging.error(f"Error in remove_clicked_pont: {e}\n{traceback.format_exc()}")
        raise

def calculate_rolling_average(df, window='10s'):
    """
    Calculates the rolling average of the 'VO2' column over a specified time window.

    This function computes a rolling average for the 'VO2' column based on a time window. 
    It fills missing values using the original data, ensuring that no information is lost.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame containing a time-indexed column named 't' and a column named 'VO2'.
    window : str, optional
        The size of the rolling window, expressed as a string compatible with pandas' time-based rolling windows (default is '10s').

    Returns
    -------
    pd.Series
        A pandas Series containing the rolling average for the 'VO2' column.
    """
    try:
        # Create a copy of the DataFrame and set 't' as the index
        df_copy = copy.deepcopy(df).set_index('t')

        # Compute rolling average and combine with original data
        df_rolling = df_copy.rolling(window=window, min_periods=1).mean().combine_first(df_copy)
        return df_rolling['VO2']
    
    except Exception as e:
        logging.error(f"Error in calculate_rolling_average: {e}\n{traceback.format_exc()}")
        raise

def get_time_range(df, start_time, end_time):
    """
    Filters a DataFrame to include rows within a specified time range.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame containing a datetime-indexed column named 't'.
    start_time : Optional[time]
        The start time for filtering. Rows with times earlier than this will be excluded. If None, no lower bound is applied.
    end_time : Optional[time]
        The end time for filtering. Rows with times later than this will be excluded. If None, no upper bound is applied.

    Returns
    -------
    pd.DataFrame
        A filtered DataFrame containing only rows where the 't' column falls within the specified time range.
    """
    try:    
        # Apply the start time filter if provided
        if start_time:
            df = df[df['t'].dt.time >= start_time]

        # Apply the end time filter if provided
        if end_time:
            df = df[df['t'].dt.time <= end_time]

        return df
    
    except Exception as e:
        logging.error(f"Error in get_time_range: {e}\n{traceback.format_exc()}")
        raise

def perform_pwlf_fit(x, y, segments):
    """
    Fits a piecewise linear model to the given data using Piecewise Linear Fitting (PWLF).

    Parameters
    ----------
    x : array-like
        The independent variable (e.g., time or input values).
    y : array-like
        The dependent variable (e.g., measurements or outputs).
    segments : int
        The number of line segments for the piecewise linear fit.

    Returns
    -------
    dict or None
        A dictionary containing the following keys if the fit is successful:
        - "slopes": Slopes of the fitted line segments.
        - "breaks": Breakpoints of the fitted model.
        - "p_values": Statistical p-values for the line segments.
        - "intercepts": Intercepts of the line segments.
        Returns `None` if the fitting fails.
    """
    try:
        # Initialize and fit the piecewise linear model
        model = pwlf.PiecewiseLinFit(x, y, seed=10)
        model.fit(segments)

        # Return the model parameters
        return {
            "slopes": model.calc_slopes(),
            "breaks": model.fit_breaks,
            "p_values": model.p_values(),
            "intercepts": model.intercepts
        }
    except Exception as e:
        logging.error(f"Error in perform_pwlf_fit: {e}\n{traceback.format_exc()}")
        raise

def modelization_time(df, search) :
    """
    Perform various time-based analyses on the DataFrame to identify key moments in the data.
    
    This function supports different types of time-based searches in the provided DataFrame:
    - 'plateau': Find the last peak in the 30-second rolling average, indicating a plateau in the data.
    - 'start': Fit a piecewise linear model on the first 3 minutes of data and determine the start point.
    - 'end': Fit a piecewise linear model on the last 3 minutes of data and determine the end point.
    - 'max': Find the time of the maximum VO2 value based on a 10-second rolling average.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame containing time ('t') and VO2 data.
    search : str
        The type of analysis to perform. Must be one of ['plateau', 'start', 'end', 'max'].

    Returns
    -------
    int or None
        If the search is successful, return the timestamp index of the found point.
        Returns `None` if the search type is unknown or an error occurs.
    """
    try:
        # Add a timestamp column in seconds
        df = df.copy()
        df["timestamp"] = pd.to_datetime(df['t'], format = "mixed").astype('int64') // 10**9

        if search == "plateau":
            # 30 seconds rolling average to identify plateau (peak)
            data_30 = calculate_rolling_average(df, '30s')
            peaks, _ = find_peaks(data_30 * -1, prominence=0.3)
            return peaks[-1] if len(peaks) > 0 else None

        elif search == "start":
            # Extract the first 3 minutes of data and perform PWLF
            subset = get_time_range(df, None, time(minute=3))
            result = perform_pwlf_fit(subset['timestamp'], subset['VO2'], 2)

            if result and subset['timestamp'].min() <= result["breaks"][1]:
                if result["slopes"][0] >= 0:
                    if result["slopes"][1] >= 0:
                        if result["p_values"][1] <= 0.05:
                            return df['timestamp'].searchsorted(result["breaks"][1], 'left')
                    if result["slopes"][1] <= 0:
                        if result["p_values"][2] <= 0.05:
                            return df['timestamp'].searchsorted(result["breaks"][2], 'left')
                elif result["slopes"][0] <= 0 and result["slopes"][1] >= 0:
                    if result["p_values"][2] <= 0.05:
                        return df['timestamp'].searchsorted(result["breaks"][2], 'left')
            return 0

        elif search == "end":
            # Extract the last 3 minutes of data and perform PWLF
            end_time = (df['t'].iloc[-1] - timedelta(minutes=3)).time()
            subset = get_time_range(df, end_time, None)
            result = perform_pwlf_fit(subset['timestamp'], subset['VO2'], 2)

            if result and result["slopes"][1] <= 0 and subset['timestamp'].max() >= result["breaks"][1]:
                if result["p_values"][1] <= 0.05:
                    return df['timestamp'].searchsorted(result["breaks"][1], 'left')
            return len(df) - 1
        
        elif search == "max":
            # 10 seconds rolling average to identify the max VO2
            data_10 = calculate_rolling_average(df)
            index_max = int(data_10.idxmax().timestamp())
            return df['timestamp'].searchsorted(index_max, 'left')
    
    except Exception as e:
        logging.error(f"Error in modelization_time: {e}\n{traceback.format_exc()}")
        raise
    
def calculate_regression(df, n, degree):
    """
    Perform regression analysis on the DataFrame's 'timestamp' and 'VO2' columns.
    
    This function can perform either a simple polynomial regression or a piecewise linear regression,
    depending on the value of `n`. If `n == 0`, it performs polynomial regression; otherwise,
    it performs piecewise linear regression.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame containing the data. Must have 'timestamp' and 'VO2' columns.
    n : int
        The number of segments for the piecewise linear regression. If `n == 0`, polynomial regression is performed.
    degree : int
        The degree of the polynomial for the polynomial regression, or the number of breaks for the piecewise linear regression.

    Returns
    -------
    np.ndarray
        The predicted values from the regression model.
    """
    try:
        # Polynomial regression (simple)
        if n == 0:  
            poly = PolynomialFeatures(degree=degree, include_bias=False)
            poly_features = poly.fit_transform(df['timestamp'].values.reshape(-1, 1))
            model = LinearRegression()
            model.fit(poly_features, df['VO2'])
            return model.predict(poly_features)
        
        # Piecewise linear regression
        else:  
            model = pwlf.PiecewiseLinFit(df['timestamp'], df['VO2'], seed=10, degree=degree)
            model.fit(n)
            return model.predict(df['timestamp'])
        
    except Exception as e:
        logging.error(f"Error in calculate_regression: {e}\n{traceback.format_exc()}")
        raise
    
def detect_outliers(df, y_predicted, std_factor, factor_up, factor_down):
    """
    Detect outliers based on the residuals between the predicted and actual 'VO2' values.
    
    The function calculates the residuals (the difference between the predicted 'VO2' and the actual 'VO2'),
    then detects outliers by comparing the residuals with a specified threshold based on the standard deviation
    and given factors for both upper and lower thresholds.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame containing the actual 'VO2' values and the corresponding timestamps 't'.
    y_predicted : np.ndarray
        The predicted 'VO2' values, typically from a regression model.
    std_factor : float
        A factor that scales the standard deviation of residuals for outlier detection.
    factor_up : float
        The scaling factor applied to the standard deviation for detecting upper outliers.
    factor_down : float
        The scaling factor applied to the standard deviation for detecting lower outliers.

    Returns
    -------
    pd.DataFrame
        DataFrame containing the timestamps ('t') and 'VO2' values of the detected outliers.
    """
    try:
        # Calculate the absolute residuals
        residuals = np.abs(y_predicted - df['VO2'].values)
        mean_residual = np.mean(residuals)
        std_residual = np.std(residuals - mean_residual)

        # Detect outliers
        outliers = df[
            (df['VO2'] > y_predicted + factor_up * std_factor * std_residual) |
            (df['VO2'] < y_predicted - factor_down * std_factor * std_residual)
        ]
        return outliers[['t', 'VO2']]
    
    except Exception as e:
        logging.error(f"Error in detect_outliers: {e}\n{traceback.format_exc()}")
        raise

def modelization_cubic_linear(df, outlier_slider_value, n):
    """
    Perform cubic or linear regression modelization on the given DataFrame and detect outliers.

    The function first converts the time column ('t') to a numeric timestamp. It then selects the appropriate 
    model parameters based on the given 'n' value and computes the regression predictions. Outliers are detected 
    based on these predictions and the specified outlier threshold.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame containing the actual 'VO2' values and the corresponding timestamps 't'.
    outlier_slider_value : float
        The threshold value for outlier detection.
    n : int
        The number of segments for piecewise linear fitting (n=0 for cubic regression, n>1 for piecewise linear regression).

    Returns
    -------
    Tuple[np.ndarray, pd.DataFrame]
        - The predicted 'VO2' values from the regression model.
        - DataFrame containing the timestamps ('t') and 'VO2' values of the detected outliers.
    """
    try:
        # Convert the time column 't' to a numeric timestamp
        df = df.copy()
        df['timestamp'] = pd.to_datetime(df['t'], format="mixed").astype('int64') // 10**9

        # Select model parameters based on the value of 'n'
        degree = 3 if n <= 1 else 1
        factor_up = 0.02
        factor_down = 0.01 if n < 2 else 0.02
        n_regression = 2 if n == 1 else n

        # Calculate the predicted 'VO2' values using regression
        y_predicted = calculate_regression(df, n_regression, degree)

        # Detect outliers based on the predicted values and specified threshold
        df_outliers = detect_outliers(df, y_predicted, outlier_slider_value, factor_up, factor_down)

        return y_predicted, df_outliers

    except Exception as e:
        logging.error(f"Error in modelization_cubic_linear: {e}\n{traceback.format_exc()}")
        raise

def compute_y_predicted(slopes, breaks, intercepts):
    """
    Compute the predicted 'y' values based on the given slopes, breaks, and intercepts.

    The function computes the predicted values by applying the regression model's 
    slopes and intercepts to the breakpoints, depending on whether the model is 
    based on 2 or 3 segments.

    Parameters
    ----------
    slopes : List[float]
        The list of slopes for each segment in the regression model.
    breaks : List[float]
        The list of breakpoints where the model changes segments.
    intercepts : List[float]
        The list of intercepts for each segment in the regression model.

    Returns
    -------
    List[float]
        A list of predicted 'y' values for each segment in the regression model.
    """
    try:
        # Two-segment model
        if len(slopes) == 2:
            return [
                intercepts[0] + slopes[0] * breaks[0],
                intercepts[1] + slopes[1] * breaks[1],
                intercepts[1] + slopes[1] * breaks[2],
            ]
        
        # Three-segment model
        elif len(slopes) == 3:
            return [
                intercepts[0] + slopes[0] * breaks[0],
                intercepts[1] + slopes[1] * breaks[1],
                intercepts[2] + slopes[2] * breaks[2],
                intercepts[2] + slopes[2] * breaks[3],
            ]
        
    except Exception as e:
        logging.error(f"Error in compute_y_predicted: {e}\n{traceback.format_exc()}")
        raise
    
def validate_RER_conditions(name, slopes, intercepts, breaks, x):
    """
    Validate the conditions for the RER model.

    This function checks if the conditions for the RER model are met, including 
    the values of the slopes, intercepts, and breaks, and ensures that the 
    breakpoint conditions align with expected ranges.

    Parameters
    ----------
    name : str
        The name of the model, used to identify if the RER model should be validated.
    slopes : List[float]
        The list of slopes for each segment in the regression model.
    intercepts : List[float]
        The list of intercepts for each segment in the regression model.
    breaks : List[float]
        The list of breakpoints for the model.
    x : List[float]
        The independent variable values used in the model.

    Returns
    -------
    tuple
        A tuple (bool, float). The boolean indicates if the RER conditions are met, 
        and the float is the value of the breakpoint where an error is detected, 
        or `None` if no error occurs.
    """
    try:
        if name == "RER":
            error_start = (slopes[0] < 0) or (intercepts[0] > 1) or (breaks[2] < 0.5 * max(x))
            if error_start:
                return False, breaks[1]
            
        return True, None

    except Exception as e:
        logging.error(f"Error in validate_RER_conditions: {e}\n{traceback.format_exc()}")
        raise

def final_modelizations(df_vo2, df_lactate = None):
    """
    Perform final modelizations for a set of physiological parameters, using piecewise linear fitting (PWLF).

    This function performs modelization for multiple parameters, including VO2, VCO2, VE, and others, 
    based on the provided VO2 and lactate data. It also validates conditions for RER and adjusts the 
    number of segments used in the piecewise linear fit as needed.

    Parameters
    ----------
    df_vo2 : pd.DataFrame
        DataFrame containing the VO2 data and corresponding parameters (e.g., 'VO2', 'VCO2', 'VE', 'FC').
    df_lactate : pd.DataFrame, optional
        DataFrame containing lactate data. If provided, the lactate values are included in the results.

    Returns
    -------
    tuple
        A tuple containing:
        - A dictionary with the results for each parameter, where each entry contains the breakpoints, 
          predicted values, slopes, intercepts, and p-values from the PWLF fitting.
        - A value indicating if an error was encountered with the RER model (if any).
    """
    try:
        # Initialization of parameter names
        list_data = ['RER', 'VCO2', 'VE', 'VE/VO2', 'VE/VCO2', 'FC', 'VO2']
        if df_lactate is not None:
            list_data.insert(5, 'Lactate')
        if 'PetCO2' in df_vo2.columns:
            list_data.insert(3, 'PetCO2')
            list_data.insert(4, 'PetO2')

        # Modeling results initialization
        results = {}
        error_RER = None

        for name in list_data:
            # Handle lactate data
            if name == "Lactate":
                results[name] = df_lactate['lactate'].tolist()
                continue

            # Choose the X axis for fitting
            x = df_vo2['FC'] if name == 'VO2' else df_vo2['VO2']
            y = df_vo2[name]

            # Choose number of segments
            n_segments = 3
            if name in ['FC', 'VO2']:
                n_segments = 2

            # Perform initial model fitting with 3 segments
            pwlf_results = perform_pwlf_fit(x, y, segments=n_segments)
            y_predicted = compute_y_predicted(pwlf_results['slopes'], pwlf_results['breaks'], pwlf_results['intercepts'])

            if name not in ['FC', 'VO2']:
                # Specific validation for RER
                is_valid, error = validate_RER_conditions(name, pwlf_results['slopes'], pwlf_results['intercepts'], pwlf_results['breaks'], x)
                if not is_valid:
                    error_RER = error
                    results[name] = [pwlf_results['breaks'], y_predicted, pwlf_results['slopes'], pwlf_results['intercepts'], pwlf_results['p_values']]
                    continue

                # Adjust to 2 segments if necessary
                if any(pwlf_results['breaks'][i] >= pwlf_results['breaks'][i + 1] - 0.4 for i in range(len(pwlf_results['breaks']) - 1)):
                    pwlf_results = perform_pwlf_fit(x, y, segments=2)
                    y_predicted = compute_y_predicted(pwlf_results['slopes'], pwlf_results['breaks'], pwlf_results['intercepts'])

            # Store the results
            results[name] = [
                pwlf_results['breaks'], 
                y_predicted, 
                pwlf_results['slopes'], 
                pwlf_results['intercepts'], 
                pwlf_results['p_values']
            ]

        return results, error_RER

    except Exception as e:
        logging.error(f"Error in final_modelizations: {e}\n{traceback.format_exc()}")
        raise
    
def _initialize_data(models):
    """
    Initializes data for modelization by extracting key information from the provided models dictionary.

    This function extracts breakpoints, slopes, p-values, and the heart rate ('FC') data from the input 
    models and prepares them for further analysis or modeling.

    Parameters
    ----------
    models : dict
        A dictionary containing model results for various physiological parameters (e.g., 'RER', 'VCO2', 'VE').
        The dictionary should include the model results in the following format:
        {
            "RER": [breakpoint, predicted_value, slope, intercept, p_value],
            ...
        }

    Returns
    -------
    dict
        A dictionary containing:
        - "breakpoints": List of breakpoints for each model parameter.
        - "slopes": List of slopes for each model parameter.
        - "pvalues": List of p-values for each model parameter.
        - "fc": Heart rate data (typically from the "FC" key in the models).
    """
    try:
        # List of default model parameters
        list_data = ["RER", "VCO2", "VE", "VE/VO2", "VE/VCO2"]

        # Add PetCO2 and PetO2 if present in models
        if "PetCO2" in models:
            list_data.extend(["PetCO2", "PetO2"])

        # Extract the required model data
        return {
            "breakpoints": [models[name][0] for name in list_data],
            "slopes": [models[name][2] for name in list_data],
            "pvalues": [models[name][-1] for name in list_data],
            "fc": models["FC"],
        }
    
    except Exception as e:
        logging.error(f"Error in _initialize_data: {e}\n{traceback.format_exc()}")
        raise

def _identify_avoid_indices(slopes):
    """
    Identifies the indices of slopes to avoid based on specific conditions.

    The function checks each slope in the input list and appends its index to the 
    `avoid` list if the slope meets certain criteria. Specifically, if the slope 
    for 'PetCO2' is non-negative or if the last value of the slope is non-positive, 
    the corresponding index is added to the `avoid` list.

    Parameters
    ----------
    slopes : list
        A list of slopes, where each element corresponds to a model parameter 
        (e.g., 'RER', 'VCO2', 'PetCO2') and the last value of each slope is checked 
        against the conditions.

    Returns
    -------
    list
        A list of indices of slopes to avoid based on the specified conditions.
    """
    try:
        avoid = []
        for i, slope in enumerate(slopes):
            if i == 5:  # PetCO2
                if slope[-1] >= 0:
                    avoid.append(i)
            elif slope[-1] <= 0:  # Other conditions
                avoid.append(i)
        return avoid
    
    except Exception as e:
        logging.error(f"Error in _identify_avoid_indices: {e}\n{traceback.format_exc()}")
        raise

def _update_threshold_choices(i, data, choiceT1, choiceT2, vo2_max, final_choice_T1, final_choice_T2):
    """
    Updates the threshold choices for T1 and T2 based on specific conditions.

    The function checks the breakpoints in the input data, comparing them with 
    certain percentages of `vo2_max`. If the breakpoints meet the criteria and 
    their corresponding p-values are below a significance threshold (0.05), 
    they are added to the respective threshold choice lists.

    Parameters
    ----------
    i : int
        The index used to access specific breakpoint data for a particular model.
    data : dict
        A dictionary containing "breakpoints" and "pvalues" lists. Each item in these lists corresponds to a model and contains breakpoint values and their associated p-values.
    choiceT1 : list
        A list to store the threshold values for T1.
    choiceT2 : list
        A list to store the threshold values for T2.
    vo2_max : float
        The maximum VO2 value, used as a reference for calculating the threshold breakpoints.
    final_choice_T1 : list
        A list to store the indices of models that contribute to the final T1 choice.
    final_choice_T2 : list
        A list to store the indices of models that contribute to the final T2 choice.

    Returns
    -------
    tuple
        A tuple containing the updated lists for `choiceT1`, `choiceT2`, `final_choice_T1`, and `final_choice_T2`.
    """
    try:
        breakpoints, pvalues = data["breakpoints"], data["pvalues"]

        # T1
        second_bp = breakpoints[i][1]
        if 0.45 * vo2_max <= second_bp < 0.70 * vo2_max and pvalues[i][1] < 0.05:
            choiceT1.append(second_bp)
            final_choice_T1.append(i)

        # T2
        if len(breakpoints[i]) == 4:  # 3-segment model
            third_bp = breakpoints[i][2]
            if 0.70 * vo2_max <= third_bp < 0.95 * vo2_max and pvalues[i][2] < 0.05:
                choiceT2.append(third_bp)
                final_choice_T2.append(i)
        else:  # 2-segment model
            if 0.70 * vo2_max <= second_bp < 0.95 * vo2_max and pvalues[i][1] < 0.05:
                choiceT2.append(second_bp)
                final_choice_T2.append(i)

        return choiceT1, choiceT2, final_choice_T1, final_choice_T2
    
    except Exception as e:
        logging.error(f"Error in _update_threshold_choices: {e}\n{traceback.format_exc()}")
        raise

def _add_lactate_thresholds(df_lactate, models, choiceT1, choiceT2):
    """
    Adds lactate thresholds to the T1 and T2 choice lists based on the computed lactate heart rate thresholds.

    The function calculates the lactate thresholds using the provided `df_lactate`, 
    then updates the T1 and T2 choices based on the VO2 model's parameters, 
    taking into account the lactate heart rate thresholds.

    Parameters
    ----------
    df_lactate : pd.DataFrame
        DataFrame containing lactate data, which is used to compute the lactate heart rate thresholds.
    models : dict
        A dictionary containing the "VO2" model with its parameters used for computing VO2 from heart rate.
    choiceT1 : list
        A list to store the threshold values for T1.
    choiceT2 : list
        A list to store the threshold values for T2.

    Returns
    -------
    tuple
        A tuple containing the updated lactate threshold data, along with the updated `choiceT1` and `choiceT2` lists.
    """
    try:
        # Compute lactate thresholds and indices
        lactate_hr_thresholds, lactate_indices, lactate_thresholds = compute_lactate(
            df_lactate
        )
        vo2_model = models["VO2"]

        # Add lactate thresholds to the T1 and T2 choices
        for i, hr in enumerate(lactate_hr_thresholds):
            if hr < vo2_model[0][1]:
                vo2 = vo2_model[3][0] + vo2_model[2][0] * hr
            else:
                vo2 = vo2_model[3][1] + vo2_model[2][1] * hr
            (choiceT1 if i == 0 else choiceT2).append(vo2)

        return [lactate_hr_thresholds, lactate_indices, lactate_thresholds], choiceT1, choiceT2
    
    except Exception as e:
        logging.error(f"Error in _add_lactate_thresholds: {e}\n{traceback.format_exc()}")
        raise

def _compute_final_threshold(choices, use_lactate, fc_model, t_type, final_choice):
    """
    Computes the final threshold values based on the provided choices, lactate usage, and the model parameters.

    The function calculates the final respiratory and heart rate thresholds based on the user's choices,
    potentially incorporating lactate data if specified. It processes the data to remove outliers and applies
    a model to compute the final heart rate.

    Parameters
    ----------
    choices : list
        List of threshold values to consider for the calculation. These can represent respiratory or lactate thresholds.
    use_lactate : bool
        Flag indicating whether to incorporate lactate data in the calculation.
    fc_model : list
        A model used to compute heart rate based on the final threshold values.
    t_type : int
        Type of threshold calculation (determines how the final value is calculated when lactate is used).
    final_choice : list
        List of user-specific choices corresponding to the threshold values.

    Returns
    -------
    dict
        A dictionary containing the following keys:
        - "respiratory" : The computed respiratory threshold.
        - "final" : The final computed threshold value.
        - "hr" : The heart rate corresponding to the final threshold value.
        - "choice" : List of indices representing the chosen threshold values.
    """
    try:
        # If no choices are provided, return NaN for all fields
        if not choices:
            return {"respiratory": np.nan, "final": np.nan, "hr": np.nan, "choice": np.nan}

        # Convert choices into a DataFrame, excluding the last value if lactate is used
        breakpoints = pd.DataFrame(choices[:-1] if use_lactate else choices)
        
        # If no valid breakpoints and lactate is used, return initial values with NaN for respiratory
        if breakpoints.shape[0] == 0 and use_lactate:
            respiratory = np.nan
            final_value = choices[0]
            new_final_choice = [7]

        else:
            # Remove outliers based on z-score if more than 2 breakpoints are present
            if breakpoints.duplicated().sum() != breakpoints.shape[0] - 1 and len(breakpoints) > 2:
                breakpoints = breakpoints[(np.abs(stats.zscore(breakpoints)) < 1).all(axis=1)]

            # Select the user's final choices from the breakpoints' indices
            new_final_choice = [final_choice[v] for v in breakpoints.index.values]

            # If lactate data is used, add the 7th value (a predefined threshold) to the choices
            if use_lactate:
                new_final_choice.append(7)

            # Calculate the respiratory threshold as the mean of the breakpoints
            respiratory = np.mean(breakpoints[0])
            final_value = respiratory

            # If lactate is used and there are more than one choice, combine respiratory and lactate thresholds
            if use_lactate and len(choices) > 1:
                if t_type == 1:
                    final_value = 0.25 * respiratory + 0.75 * choices[-1]
                else:
                    final_value = 0.5 * (respiratory + choices[-1])

        # Calculate the heart rate based on the final threshold value using the provided model
        if final_value < fc_model[0][1]:
            hr = fc_model[3][0] + fc_model[2][0] * final_value
        else:
            hr = fc_model[3][1] + fc_model[2][1] * final_value

        return {"respiratory": respiratory, "final": final_value, "hr": hr, "choice": new_final_choice}
    
    except Exception as e:
        logging.error(f"Error in _compute_final_threshold: {e}\n{traceback.format_exc()}")
        raise

def computation(df_lactate, models, test_id, use_lactate):
    """
    Computes the threshold values for respiratory and lactate thresholds, 
    as well as heart rate based on the provided models and test data.

    The function handles the threshold choices for T1 and T2, including lactate thresholds if required, 
    and calculates the corresponding final respiratory thresholds, lactate thresholds, and heart rate values.

    Parameters
    ----------
    df_lactate : pd.DataFrame
        DataFrame containing the lactate values for the test session.
    models : dict
        Dictionary containing the models for respiratory and lactate data.
    test_id : int
        The ID of the test used to fetch the VO2 max value.
    use_lactate : bool
        Flag to determine if lactate data should be included in the computation.

    Returns
    -------
    tuple
        A tuple containing the following elements:
        - Respiratory threshold for T1 and T2
        - Lactate threshold values (if applicable)
        - Final threshold values for T1 and T2
        - Heart rate corresponding to the final threshold values
        - Indices of the chosen thresholds
    """
    try:
        # Fetch the session and VO2 max for the given test ID
        session = next(get_db())
        vo2_max = get_by_id(session, Test, test_id).vo2_max

        # Initialize constants and data structures
        respiratory_data = _initialize_data(models)
        choiceT1, choiceT2, final_choice_T1, final_choice_T2 = [], [], [], []
        avoid_indices = _identify_avoid_indices(respiratory_data["slopes"])

        # Calculate the threshold choices for T1 and T2
        for i in range(len(respiratory_data["breakpoints"])):
            if i in avoid_indices:
                continue
            choiceT1, choiceT2, final_choice_T1, final_choice_T2 = _update_threshold_choices(
                i, respiratory_data, choiceT1, choiceT2, vo2_max, final_choice_T1, final_choice_T2
            )

        # Add lactate thresholds if needed
        if use_lactate:
            lactate_results, choiceT1, choiceT2 = _add_lactate_thresholds(
                df_lactate, models, choiceT1, choiceT2
            )
        else:
            lactate_results = [[], [], []]

        # Compute the final thresholds for T1 and T2
        t1_results = _compute_final_threshold(
            choiceT1, use_lactate, respiratory_data["fc"], 1, final_choice_T1
        )
        t2_results = _compute_final_threshold(
            choiceT2, use_lactate, respiratory_data["fc"], 2, final_choice_T2
        )

        return (
            t1_results["respiratory"],
            t2_results["respiratory"],
            lactate_results[0],
            lactate_results[1],
            lactate_results[2],
            t1_results["final"],
            t2_results["final"],
            t1_results["hr"],
            t2_results["hr"],
            t1_results["choice"],
            t2_results["choice"],
        )
    
    except Exception as e:
        logging.error(f"Error in computation: {e}, {traceback.format_exc()}")
        raise

def _find_threshold_index(df, diff_threshold):
    """
    Find the first index where lactate_diff exceeds the given threshold.
    
    This function checks the 'lactate_diff' column of the provided DataFrame and
    returns the index of the first occurrence where the value exceeds the specified
    threshold. If no such index is found, it returns None.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame containing the 'lactate_diff' column with lactate difference values.
    diff_threshold : float
        The threshold value above which the lactate difference is considered significant.

    Returns
    -------
    int or None
        The index of the first row where 'lactate_diff' exceeds the threshold,
        or None if no such value is found.
    """
    try:
        # Find indices where lactate_diff exceeds the threshold
        indices = df.index[df['lactate_diff'] > diff_threshold].to_list()

        # Return the first index if available, else None
        if not indices:
            return None
        
        return indices[0]
    
    except Exception as e:
        logging.error(f"Error in _find_threshold_index: {e}\n{traceback.format_exc()}")
        raise

def _compute_first_threshold(df, index_first, index_min):
    """
    Compute the first lactate threshold and associated data.
    
    This function calculates the first lactate threshold, which involves determining
    the heart rate (hr), lactate value, and the index of the test for the threshold 
    based on the lactate difference. The values are computed either from the previous 
    point or as the average of the two surrounding points.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame containing the columns 'lactate_diff', 'fc' (heart rate), and 'lactate'.
    index_first : int
        The index of the first point where the lactate difference exceeds the threshold.
    index_min : int
        The minimum index value used in calculating the threshold.

    Returns
    -------
    tuple
        A tuple containing:
        - hr : list of heart rate(s) corresponding to the lactate threshold.
        - index_test : list containing the calculated index of the threshold.
        - lactate : list of lactate values corresponding to the threshold.
    """
    try:
        # Check if lactate difference at the first index exceeds the threshold
        if df['lactate_diff'].iloc[index_first] >= 0.3:
            hr = [df['fc'].iloc[index_first - 1]]
            index_test = [index_first + index_min - 1]
            lactate = [df['lactate'].iloc[index_first - 1]]
        else:
            hr = [np.mean([df['fc'].iloc[index_first - 1], df['fc'].iloc[index_first]])]
            index_test = [index_min + index_first - 0.5]
            lactate = [np.mean([df['lactate'].iloc[index_first - 1], df['lactate'].iloc[index_first]])]

        return hr, index_test, lactate
    
    except Exception as e:
        logging.error(f"Error in _compute_first_threshold: {e}\n{traceback.format_exc()}")
        raise

def _compute_second_threshold(df, index_second, index_first, index_min):
    """
    Compute the second lactate threshold iteratively.
    
    This function computes the second lactate threshold by iterating over the dataframe
    to find a significant change in the lactate difference or detect a constant value.
    The lactate threshold, heart rate, and test index are calculated and returned.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame containing the columns 'lactate_diff', 'fc' (heart rate), and 'lactate'.
    index_second : int
        The index of the second point in the threshold computation.
    index_first : int
        The index of the first point used to compute the threshold.
    index_min : int
        The minimum index value used in calculating the second threshold.

    Returns
    -------
    tuple
        A tuple containing:
        - hr_thresholds : list of heart rate(s) corresponding to the second lactate threshold.
        - index_test_thresholds : list containing the calculated index of the second threshold.
        - lactate_thresholds : list of lactate values corresponding to the second threshold.
    """
    try:
        df = df.iloc[index_second - 1:].reset_index(drop=True)
        hr_thresholds = []
        index_test_thresholds = []
        lactate_thresholds = []

        major_index = index_first + index_min + index_second - 1
        second_threshold = False

        # Temporary save the first value as the threshold
        temp_hr_threshold = df['fc'].iloc[0]
        temp_index_test_threshold = df.shape[0]-1
        temp_lactate_threshold = df['lactate'].iloc[0]

        # Only one value left in the dataframe
        if df.shape[0] <= 1 :
            hr_thresholds.append(temp_hr_threshold)
            index_test_thresholds.append(temp_index_test_threshold)
            lactate_thresholds.append(temp_lactate_threshold)

        # Multiple values left in the dataframe
        else :
            # Scan value by value
            while not second_threshold and df.shape[0] > 1:

                # Get the previous index if the next gap if higher than 1 or the values are constant
                if df['lactate_diff'].iloc[2] > 1 or df['lactate_diff'].iloc[2:].nunique() == 1:
                    hr_thresholds.append(df['fc'].iloc[0])
                    index_test_thresholds.append(major_index)
                    lactate_thresholds.append(df['lactate'].iloc[0])
                    second_threshold = True
                else:
                    df = df.iloc[1:].reset_index(drop=True)
                    major_index += 1

            # If no second threshold value has been found
            if not second_threshold:
                hr_thresholds.append(temp_hr_threshold)
                index_test_thresholds.append(temp_index_test_threshold)
                lactate_thresholds.append(temp_lactate_threshold)

        return hr_thresholds, index_test_thresholds, lactate_thresholds
    
    except Exception as e:
        logging.error(f"Error in _compute_second_threshold: {e}\n{traceback.format_exc()}")
        raise

def compute_lactate(df):
    """
    Compute lactate thresholds based on the provided DataFrame containing lactate and heart rate (fc) data.
    
    This function calculates two lactate thresholds based on the difference in lactate values. 
    It identifies the minimum lactate point, computes the first and second lactate thresholds,
    and returns the corresponding heart rate, lactate values, and index of the thresholds.
    
    Parameters
    ----------
    df : pd.DataFrame
        DataFrame containing 'lactate' and 'fc' columns.

    Returns
    -------
    tuple
        A tuple containing:
        - hr_thresholds : List of heart rates corresponding to the lactate thresholds.
        - index_test_thresholds : List of indices where the lactate thresholds were identified.
        - lactate_thresholds : List of lactate values at the identified thresholds.
    """
    try:
        # Validate the required columns
        required_columns = {'lactate', 'fc'}
        if not required_columns.issubset(df.columns):
            raise ValueError(f"Le DataFrame doit contenir les colonnes suivantes : {required_columns}")

        # Copy data and calculate lactate differences
        df_lactate = df.copy()
        df_lactate['lactate_diff'] = df_lactate['lactate'].diff()

        # Step 1: Identify the starting point after the last minimum
        if -1 in np.sign(df_lactate['lactate_diff'].values):
            index_min_lactate = df_lactate[df_lactate['lactate_diff'] < 0].index.max()
        else:
            index_min_lactate = df_lactate['lactate'][::-1].idxmin()

        df_lactate = df_lactate.iloc[index_min_lactate:].reset_index(drop=True)

        # Step 2: Calculate the first threshold
        index_first_threshold = _find_threshold_index(df_lactate, diff_threshold=0.1)
        hr_thresholds, index_test_thresholds, lactate_thresholds = _compute_first_threshold(
            df_lactate, index_first_threshold, index_min_lactate
        )

        # Step 3: Calculate the second threshold
        df_lactate = df_lactate.iloc[index_first_threshold:].reset_index(drop=True)
        index_second_threshold = _find_threshold_index(df_lactate, diff_threshold=1)

        if not index_second_threshold:
            # If no second threshold is found, use the maximum lactate difference
            index_second_threshold = df_lactate['diff'].idxmax()
            second_threshold_data = [
                hr_thresholds.append(df_lactate['fc'].iloc[index_second_threshold]),
                index_test_thresholds.append(index_first_threshold + index_min_lactate + index_second_threshold - 1),
                lactate_thresholds.append(df_lactate['lactate'].iloc[index_second_threshold])
            ]
        else:
            # Calculate the second threshold if the index is found
            second_threshold_data = _compute_second_threshold(
                df_lactate, index_second_threshold, index_first_threshold, index_min_lactate
            )

        # Add the results of the second threshold
        hr_thresholds.extend(second_threshold_data[0])
        index_test_thresholds.extend(second_threshold_data[1])
        lactate_thresholds.extend(second_threshold_data[2])

        return hr_thresholds, index_test_thresholds, lactate_thresholds

    except Exception as e:
        logging.error(f"Error in compute_lactate: {e}\n{traceback.format_exc()}")
        raise
    
def compute_label_checkbox(respiratory_threshold, hr_curve):
    """
    Compute the heart rate based on the respiratory threshold and heart rate curve.

    This function calculates the heart rate corresponding to a given respiratory threshold using
    the provided heart rate curve data. It checks if the respiratory threshold is below or above
    the value specified in the first part of the curve and applies the appropriate formula to calculate
    the heart rate.

    Parameters
    ----------
    respiratory_threshold : float
        The value of the respiratory threshold to compute the corresponding heart rate.
    hr_curve : list of lists
        A list containing the heart rate curve data. The first element is a list containing the
        threshold values, the second is the slope values for the first section of the curve,
        and the third contains the slope values for the second section.

    Returns
    -------
    str
        A string representing the computed heart rate, rounded to the nearest integer, followed
        by "bpm" (e.g., "120 bpm").
    """
    try:
        # Check if the respiratory threshold is below the first value in the HR curve
        if respiratory_threshold < hr_curve[0][1] :
            hr = hr_curve[1][0] + hr_curve[2][0]*respiratory_threshold
        else :
            hr = hr_curve[1][1] + hr_curve[2][1]*respiratory_threshold

        # Round the heart rate to the nearest integer and format it
        hr = f"{round(hr)} bpm"

        return hr
    
    except Exception as e:
        logging.error(f"Error in compute_label_checkbox: {e}\n{traceback.format_exc()}")
        raise

def interpolate(value, series):
    """
    Interpolate a value within a given series (e.g., a pandas Series).
    
    This function performs linear interpolation between two points in a series, based on the 
    provided value. The interpolation is done between the base index and the next index. If the 
    value is at the last index or out of bounds, it simply returns the corresponding series value.
    
    Parameters
    ----------
    value : float
        The value at which to interpolate.
    series : pandas.Series
        A pandas Series that holds the data for interpolation.
        
    Returns
    -------
    float
        The interpolated value corresponding to the provided value.
    """
    try:
        # Ensure the series has at least two elements for interpolation
        base = math.floor(value)

        # Base index and interpolation check
        if base + 1 < len(series):
            diff = value - base
            return series.iloc[base] + diff * (series.iloc[base + 1] - series.iloc[base])
        
        # If the value is out of bounds, return the last element of the series
        return series.iloc[base]
    
    except Exception as e:
        logging.error(f"Error in interpolate: {e}\n{traceback.format_exc()}")
        raise

def update_threshold_computation(value, curves, results, threshold_type, threshold_level, df_lactate):
    """
    Update the computation of thresholds based on the provided value, type, and level.

    This function updates the threshold values (lactate or respiratory) based on the provided 
    input value and updates the results dictionary with computed lactate, heart rate (HR), 
    and final threshold values. The function supports both lactate and respiratory threshold types.
    
    Parameters
    ----------
    value : float
        The new threshold value to be computed (e.g., lactate or respiratory threshold).
    curves : dict
        A dictionary containing the FC, VO2, and possibly other curves used in the computation.
    results : list
        A list of results that stores the computed thresholds and related values. It gets updated 
        with the new threshold values, HR values, and labels.
    threshold_type : str
        The type of threshold being updated. Either 'lactate' or 'respiratory'.
    threshold_level : int
        The index of the threshold level being updated (0 or 1).
    df_lactate : pandas.DataFrame or None
        A DataFrame containing lactate data. If not provided, the function computes only the respiratory thresholds.
    
    Returns
    -------
    tuple
        A tuple containing the updated results list and the label for the threshold.
    """
    try:
        # Get the position in the FC curve
        index_fc_curve = 0 if value < curves['FC'][0][1] else 1

        if df_lactate is not None:

            # Handle lactate threshold computation
            if threshold_type == "lactate":

                # Update the new position value
                results[3][threshold_level] = value

                # Compute the new lactate and hr threshold value
                results[2][threshold_level] = interpolate(value, df_lactate['fc'])
                results[4][threshold_level] = interpolate(value, df_lactate['lactate'])
                label = f"{round(results[2][threshold_level])} bpm"

            # Handle respiratory threshold computation
            elif threshold_type == "respiratory":
                
                # Update the respiratory threshold
                results[threshold_level] = value

                # Update the threshold label
                label = f"{round(curves['FC'][1][index_fc_curve] + (curves['FC'][2][index_fc_curve]*value))} bpm"

            # Update the lactate threshold
            index_vo2_curve = 0 if results[2][threshold_level] < curves['VO2'][0][1] else 1
            lactate_threshold = curves['VO2'][1][index_vo2_curve] + curves['VO2'][2][index_vo2_curve] * results[2][threshold_level]

            # Compute the final threshold value
            if threshold_level == 0:
                results[5] = results[0]*0.25 + lactate_threshold*0.75
            elif threshold_level == 1:
                results[6] = (results[1] + lactate_threshold)*0.5
        
        else:

            # Compute the final threshold value and update the respiratory threshold
            results[threshold_level], results[threshold_level+5] = value, value

            # Update the threshold label
            label = f"{round(curves['FC'][1][index_fc_curve] + (curves['FC'][2][index_fc_curve]*value))} bpm"

        # Update the final threshold HR value 
        index_final_fc_curve = 0 if results[threshold_level+5] < curves['FC'][0][1] else 1
        results[threshold_level+7] = curves['FC'][1][index_final_fc_curve] + curves['FC'][2][index_final_fc_curve]*results[threshold_level+5]

        return results, label
    
    except Exception as e:
        logging.error(f"Error in update_threshold_computation: {e}\n{traceback.format_exc()}")
        raise

def compute_threshold_position(next_fc, threshold, prev_fc, prev_index):
    """
    Compute the position of a threshold value on a given curve.

    This function calculates the relative position of a threshold between two points
    on the curve, given the threshold value and the previous and next data points.
    
    The formula is based on linear interpolation between the two points: the previous
    point (`prev_fc`) and the next point (`next_fc`).

    Parameters
    ----------
    next_fc : float
        The value of the next data point in the curve (after `prev_fc`).
    threshold : float
        The threshold value to find the corresponding position for on the curve.
    prev_fc : float
        The value of the previous data point in the curve (before `next_fc`).
    prev_index : int
        The index of the previous data point in the curve. This is used to scale the result.

    Returns
    -------
    float
        The relative position of the threshold between `prev_fc` and `next_fc`, adjusted 
        by `prev_index`. The position is calculated using linear interpolation.
    """
    try:
        return ((threshold - prev_fc) / (next_fc - prev_fc)) + prev_index
    
    except Exception as e:
        logging.error(f"Error in compute_threshold_position: {e}\n{traceback.format_exc()}")
        raise

def subcompute_levels(value, curve):
    """
    Compute the level for a given value based on a curve.

    This function calculates the corresponding value on a curve based on the provided
    input value. The curve is assumed to be in the form of a 3-tuple, where:
    - `curve[0]` contains the x-values (independent variable),
    - `curve[1]` contains the y-values (dependent variable),
    - `curve[2]` contains the slope coefficients (for linear interpolation).

    The function uses linear interpolation to find the corresponding y-value for the
    given input `value`.

    Parameters
    ----------
    value : float
        The input value to compute the corresponding level for.
    curve : tuple
        A tuple containing three lists:
        - `curve[0]` (list of x-values),
        - `curve[1]` (list of y-values),
        - `curve[2]` (list of slope coefficients for linear interpolation).

    Returns
    -------
    float
        The computed y-value for the given input `value` on the curve.
    """
    try:
        if value < curve[0][0]:
            return curve[1][0] + curve[2][0] * value
        if value >= curve[0][-1]:
            return curve[1][-1] + curve[2][-1] * value
        for v in range(1, len(curve[0])):
            if curve[0][v-1] <= value < curve[0][v]:
                return curve[1][v-1] + curve[2][v-1] * value
            
    except Exception as e:
        logging.error(f"Error in subcompute_levels: {e}\n{traceback.format_exc()}")
        raise
        
def compute_x_vo2_levels(s1, s2, df_lactate, curves):
    """
    Compute the VO2 levels and corresponding x-values for the lactate threshold levels.

    This function computes the x-values (positions) where the thresholds s1 and s2
    occur in the lactate curve, as well as the corresponding VO2 values for each
    lactate data point in `df_lactate`.

    Parameters
    ----------
    s1 : float
        The first threshold value for which the x-position will be computed.
    s2 : float
        The second threshold value for which the x-position will be computed.
    df_lactate : pandas.DataFrame
        A DataFrame containing lactate data, specifically with 'fc' (heart rate) values.
    curves : dict
        A dictionary containing curve data for different parameters. The 'VO2' curve is expected
        to have three elements: x-values, y-values, and slope coefficients.

    Returns
    -------
    tuple
        A tuple containing:
        - A list of x-values (positions) where thresholds s1 and s2 occur in the lactate data.
        - A list of VO2 levels corresponding to the heart rate values in `df_lactate`.
    """
    try:
        # Initialization of levels
        x_levels = []
        vo2_levels = []

        # Compute the x-values and VO2 levels
        for f in range(len(df_lactate['fc'])):
            if f > 0:
                prev_fc = df_lactate['fc'].iloc[f-1]
                next_fc = df_lactate['fc'].iloc[f]
                prev_index = df_lactate.index[f-1]

                if prev_fc < s1 <= next_fc:
                    x_levels.append(compute_threshold_position(next_fc, s1, prev_fc, prev_index))
                if prev_fc <= s2 < next_fc:
                    x_levels.append(compute_threshold_position(next_fc, s2, prev_fc, prev_index))

            # Calculate the VO2 level for each heart rate
            fc_value = df_lactate['fc'].iloc[f]
            vo2_levels.append(subcompute_levels(fc_value, curves['VO2']))

        return x_levels, vo2_levels
    
    except Exception as e:
        logging.error(f"Error in compute_x_vo2_levels: {e}\n{traceback.format_exc()}")
        raise

def compute_levels(results, df_lactate, curves, test_id):
    """
    Compute various physiological levels based on lactate data, curve parameters, and test information.

    This function calculates the x-levels, VO2 levels, and various other parameters such as VE, VCO2,
    VE/VO2, RER, glucidic and lipidic contributions, and normalizes VO2 and VCO2 levels by body weight.

    Parameters
    ----------
    results : list
        A list containing the computed threshold values including s1 and s2 for lactate thresholds.
    
    df_lactate : pandas.DataFrame
        A DataFrame containing lactate data with heart rate ('fc') and lactate ('lactate') values.
    
    curves : dict
        A dictionary containing curve data for VO2, VCO2, VE, VE/VO2, and RER parameters.
    
    test_id : int
        The ID of the test from the database to retrieve the body weight for normalization of VO2/VCO2.

    Returns
    -------
    dict
        A dictionary with the following keys:
        - 'x': x-levels for lactate thresholds.
        - 'vo2': VO2 levels normalized by body weight.
        - 've': Ventilation levels.
        - 'vco2': VCO2 levels normalized by body weight.
        - 'vevo2': Ventilation per oxygen consumption ratio.
        - 'rer': Respiratory exchange ratio.
        - 'glu': Glucidic contribution.
        - 'lip': Lipidic contribution.  
    """
    try:
        # Compute the x and VO2 levels using the lactate data
        x_levels, vo2_levels = compute_x_vo2_levels(results[7], results[8], df_lactate, curves)

        # Get the VE, VCO2, VE/VO2, and RER values for each level
        ve_levels = [subcompute_levels(level, curves['VE']) for level in vo2_levels]
        vco2_levels = [subcompute_levels(level, curves['VCO2']) for level in vo2_levels]
        vevo2_levels = [subcompute_levels(level, curves['VE/VO2']) for level in vo2_levels]
        rer_levels = [subcompute_levels(level, curves['RER']) for level in vo2_levels]
        
        # Calculate glucidic and lipidic contributions, ensuring non-negative values
        glu_levels = [max(0, 4.585 * vco2 - 3.2255 * vo2) for vco2, vo2 in zip(vco2_levels, vo2_levels)]
        lip_levels = [max(0, 1.6946 * vo2 - 1.7012 * vco2) for vco2, vo2 in zip(vco2_levels, vo2_levels)]

        # Normalize VO2 and VCO2 levels by body weight
        session = next(get_db())
        test = get_by_id(session, Test, test_id)
        vo2_levels = [v * 1000/test.weight for v in vo2_levels]
        vco2_levels = [v * 1000/test.weight for v in vco2_levels]

        # Return the computed values in a dictionary
        return {
            'x': x_levels, 
            'vo2': vo2_levels, 
            've': ve_levels, 
            'vco2': vco2_levels, 
            'vevo2': vevo2_levels, 
            'rer': rer_levels,
            'glu': glu_levels,
            'lip': lip_levels
        }
    
    except Exception as e:
        logging.error(f"Error in compute_levels: {e}\n{traceback.format_exc()}")
        raise

def compute_value(index, threshold, curve):
    """
    Compute a value from a given threshold using the parameters from the curve.

    This function calculates the value by using the formula:
    value = curve[1][index] + curve[2][index] * threshold
    
    Parameters
    ----------
    index : int
        The index of the value to retrieve from the curve.
    
    threshold : float
        The threshold value to apply in the calculation.
    
    curve : list
        A list containing curve data with the following structure:
        - curve[0]: X-axis values (not used in this function)
        - curve[1]: Y-axis values (for the given index)
        - curve[2]: Slope values for the calculation.

    Returns
    -------
    float
        The computed value from the curve based on the provided threshold.
    """
    try:
        # Calculate the value
        return curve[1][index] + curve[2][index] * threshold
    
    except Exception as e:
        logging.error(f"Error in compute_value: {e}\n{traceback.format_exc()}")
        raise
    
def compute_ve_vco2_values(results, curve, vo2_max) :
    """
    Compute VE (Ventilatory Equivalent) and VCO2 (Carbon Dioxide Output) values for given thresholds and VO2 max.

    This function processes the threshold values (results[5] and results[6]) by evaluating them against 
    a provided curve. It also computes a final value based on VO2 max, then returns the list of computed values, 
    sorted in ascending order.

    Parameters
    ----------
    results : list
        A list containing threshold values. Specifically, results[5] and results[6] are used for threshold computation.
    
    curve : list
        A list of lists representing the curve data. The structure of `curve` is expected to be:
        - curve[0]: The independent variable (e.g., VO2 values).
        - curve[1]: The dependent variable (e.g., VE or VCO2 values).
        - curve[2]: The slope values for the curve.

    vo2_max : float
        The VO2 max value used for the final computation.

    Returns
    -------
    list
        A sorted list of computed VE and VCO2 values based on the thresholds and VO2 max.
    """
    try:
        # Initialize the result array
        result = []

        # Process each threshold (results[5] and results[6])
        thresholds = [results[5], results[6]]
        for threshold in thresholds:
            # Below curve range
            if threshold < curve[0][0]:  
                result.append(compute_value(0, threshold, curve))
            # Above curve range
            elif threshold >= curve[0][-1]:  
                result.append(compute_value(-1, threshold, curve))
            # Within curve range
            else:  
                for v in range(1, len(curve[0])):
                    if curve[0][v-1] <= threshold < curve[0][v]:
                        result.append(compute_value(v-1, threshold, curve))
                        break

        # Add the value computed with VO2 max
        result.append(compute_value(-1, vo2_max, curve))

        # Sort the results
        result.sort()

        return result
    
    except Exception as e:
        logging.error(f"Error in compute_ve_vco2_values: {e}\n{traceback.format_exc()}")
        raise

def compute_values(curves, results, vo2_max):
    """
    Compute the values for VE, VCO2, glucose (glu), and lipids (lip) based on the provided curves and results.

    This function processes the VE and VCO2 values using the given curves and computes the glucose and lipid 
    values based on specified coefficients. It also handles the final calculation for the VO2 max.

    Parameters
    ----------
    curves : dict
        A dictionary containing the curves for VE and VCO2, where each curve is expected to be a list of three elements:
        - curve[0]: Independent variable (e.g., VO2 values).
        - curve[1]: Dependent variable (e.g., VE or VCO2 values).
        - curve[2]: Slope values for the curve.

    results : list
        A list containing the threshold values and other results used in the calculation. Specifically, it contains 
        at least the lactate threshold results and the VO2 thresholds at results[5], results[6], and other necessary 
        values.

    vo2_max : float
        The VO2 max value used for the final computations.

    Returns
    -------
    dict
        A dictionary with the following keys and computed values:
        - 've': A list of computed VE values.
        - 'vco2': A list of computed VCO2 values.
        - 'glu': A list of glucose values.
        - 'lip': A list of lipid values.
    """
    try:
        # Initialize the values
        values = {} 

        # Compute the VE and VCO2 values
        values['ve'] = compute_ve_vco2_values(results, curves['VE'], vo2_max)
        values['vco2'] = compute_ve_vco2_values(results, curves['VCO2'], vo2_max)

        # Define coefficients for glucose and lipid calculation
        COEF_GLU_VCO2 = 4.585
        COEF_GLU_TEMP = 3.2255
        COEF_LIP_TEMP = 1.6946
        COEF_LIP_VCO2 = 1.7012

        # Compute glucose (glu) and lipid (lip) values based on coefficients
        glu = [max(0, COEF_GLU_VCO2 * values['vco2'][i] - COEF_GLU_TEMP * results[i+5]) for i in range(2)]
        lip = [max(0, COEF_LIP_TEMP * results[i+5] - COEF_LIP_VCO2 * values['vco2'][i]) for i in range(2)]

        # Additional calculations for glucose and lipid values with VO2 max
        glu.append(max(0, COEF_GLU_VCO2 * values['vco2'][2] - COEF_GLU_TEMP * vo2_max))
        lip.append(max(0, COEF_LIP_TEMP * vo2_max - COEF_LIP_VCO2 * values['vco2'][2]))

        # Add the computed values to the dictionary
        values['glu'] = glu
        values['lip'] = lip

        return values
    
    except Exception as e:
        logging.error(f"Error in compute_values: {e}\n{traceback.format_exc()}")
        raise

def intersect(h, s, b, inter, y) :
    """
    Method to determine the intersection points between the threshold values and the modelization curve.
    
    Parameters
    ----------
    h : float
        VO2 ordinate value for the intersection. The value at which we want to find the intersection on the y-axis.
    s : array-like
        Array of slopes for the three segments of the modelization.
    b : array-like
        Array of breakpoints for the three segments of the modelization.
    inter : array-like
        Array of intercepts for the three segments of the modelization.
    y : array-like
        Array of y values for the beginning of the three segments of the modelization.
    
    Returns
    -------
    list or None
        List of intersection points `[x, y]` where the x is the position of the intersection on the x-axis and 
        y is the corresponding value on the model curve. Returns None if no intersection is found.
    """
    try:
        # Initialization of the intersection
        xi = None
        yi = None

        # Iterate over the three pieces of the modelization
        n_pieces = len(s)
        for i in range(n_pieces):

            # Find x and y for the current segment
            xitemp = (h - inter[i]) / s[i]
            yitemp = h

            # Validate the intersection for the current piece
            if i < n_pieces - 1:  # Not the last segment
                if b[i] <= xitemp <= b[i + 1] and y[i] <= yitemp <= y[i + 1]:
                    xi = xitemp
                    yi = yitemp
                    break
            else:  # Last segment
                if b[i] <= xitemp and y[i] <= yitemp:
                    xi = xitemp
                    yi = yitemp
                    break

        # Return the intersection point (or None if not found)
        return [xi, yi] if xi is not None and yi is not None else None
    
    except Exception as e:
        logging.error(f"Error in intersect: {e}\n{traceback.format_exc()}")
        raise

def find_intersection(x1, y1, x2, y2):
    """
    Finds the x-coordinate where the line between two points (x1, y1) and (x2, y2)
    intersects the horizontal line y = 0.

    Parameters
    ----------
    x1 : float
        The x-coordinate of the first point.
    y1 : float
        The y-coordinate of the first point.
    x2 : float
        The x-coordinate of the second point.
    y2 : float
        The y-coordinate of the second point.

    Returns
    -------
    float
        The x-coordinate of the intersection with the y = 0 line.
    """
    try:
        # Calculate the x-coordinate of the intersection
        return x1 + (x2 - x1) * abs(y1) / abs(y1 - y2)
    
    except Exception as e:
        logging.error(f"Error in find_intersection: {e}\n{traceback.format_exc()}")
        raise

def integral(x, d1, d2, y_1, y_2) :
    """ 
    Method to compute the integral values of the VO2 plateau 
    
    Parameters
    ----------
    x : array
        Array of indexes of the data
    d1 : lambda
        Function of positive difference between the cubic modelization of the plateau and the linear modelization between the two thresholds
    d2 : lambda
        Function of negative difference between the cubic modelization of the plateau and the linear modelization between the two thresholds
    y_1 : array
        Array of ordinate values for the cubic modelization of the plateau
    y_2 : array
        Array of ordinate values for the linear modelization between the two thresholds

    Returns
    -------
    array
        Array of time boundaries under and above and area values under and above the linear trend
    """
    try:
        # Initialization of the position of the crosses around the linear trend
        b = 0
        p = 0

        # Initialization of the time boundaries arrays
        if y_1.iloc[0] <= y_2.iloc[0] :
            esub = [[x.iloc[0]]]
            esup = [[]]

        else :
            esub = [[]]
            esup = [[x.iloc[0]]]

        # Iterate over the data indexes : determine the time boundaries of the sub and sup areas
        for i in range(1, len(x)) :

            # Check if an intersection happened between two points
            if (y_1.iloc[i-1] < y_2.iloc[i-1] and y_1.iloc[i] > y_2.iloc[i]) or (y_1.iloc[i-1] > y_2.iloc[i-1] and y_1.iloc[i] < y_2.iloc[i]) :
                xi = find_intersection(x.iloc[i-1], y_1.iloc[i-1] - y_2.iloc[i-1], x.iloc[i], y_1.iloc[i] - y_2.iloc[i])
                esub[b].append(xi)
                esup[p].append(xi)
            # Check if it is an intersection point
            elif y_1.iloc[i] == y_2.iloc[i] :
                esub[b].append(x.iloc[i])
                esup[p].append(x.iloc[i])

            # Add a new range if the past one has been completed
            if len(esub[b]) == 2 :
                esub.append([])
                b += 1
            if len(esup[p]) == 2 :
                esup.append([])
                p += 1

        # Finalize the time boundaries array with the last index
        if y_1.iloc[-1] < y_2.iloc[-1] :
            esub[b].append(x.iloc[-1])
        else :
            esup[p].append(x.iloc[-1])

        # Remove empty arrays
        esub = [arr for arr in esub if arr]
        esup = [arr for arr in esup if arr]

        # Calculate the integrals with the associated time boundaries
        integ_sub = []
        integ_sup = []
        for e in esub :
            integ_sub.append([round(abs(v)) for v in quad(d2, e[0], e[1])])
        for s in esup :
            integ_sup.append([round(v) for v in quad(d1, s[0], s[1])])

        return [esub, esup, integ_sub, integ_sup]
    
    except Exception as e:
        logging.error(f"Error in integral: {e}\n{traceback.format_exc()}")
        raise
            
def compute_mse(df):
    """ 
    Method to compute the Mean Squared Error (MSE) for linear and polynomial (degree 2) modelizations
    
    Parameters
    ----------
    df : DataFrame
        DataFrame containing the 'timestamp' and 'VO2' columns for the fitting
    
    Returns
    -------
    tuple
        A tuple where the first value indicates the model (0 for linear, 1 for polynomial) with the smallest MSE,
        and the second value is a list containing the coefficients and residuals for both models.
    """
    try:
        # Compute a linear and a polynomial of degree 2 modelization of the plateau
        z = []

        # Fit a linear model (degree 1)
        coeffs_linear, residuals_linear, _, _, _ = np.polyfit(df['timestamp'], df['VO2'], 1, full=True)
        z.append((coeffs_linear, residuals_linear))

        # Fit a polynomial of degree 2
        coeffs_poly2, residuals_poly2, _, _, _ = np.polyfit(df['timestamp'], df['VO2'], 2, full=True)
        z.append((coeffs_poly2, residuals_poly2))

        # Find the most accurate one: the smaller residuals value
        # Residuals for each model are stored in z[0][1] for the linear model and z[1][1] for the polynomial model
        # In case residuals are empty, calculate them manually
        if len(z[0][1]) == 0:
            y_fit_linear = np.polyval(z[0][0], df['timestamp'])
            residuals_linear = df['VO2'] - y_fit_linear
            mse_linear = np.mean(residuals_linear**2)
        else:
            mse_linear = z[0][1][0]  # The residual sum of squares (RSS) from np.polyfit

        if len(z[1][1]) == 0:
            y_fit_poly2 = np.polyval(z[1][0], df['timestamp'])
            residuals_poly2 = df['VO2'] - y_fit_poly2
            mse_poly2 = np.mean(residuals_poly2**2)
        else:
            mse_poly2 = z[1][1][0]  # The residual sum of squares (RSS) from np.polyfit

        # Find the model with the smaller MSE (mean squared error)
        min_mse = np.argmin([mse_linear, mse_poly2])

        return min_mse, z
    
    except Exception as e:
        logging.error(f"Error in compute_mse: {e}\n{traceback.format_exc()}")
        raise

def compute_plateau(test_data, vo2_results) :
    """ 
    Main part of the code for the plateau analysis

    It computes the results and the figures.

    Parameters
    ----------
    test_data : Test
        Test object in order to retrieve the data
    vo2_results : list
        A list containing the threshold VO2 values to be used in the analysis

    Returns
    -------
    tuple
        A tuple containing the plot figure and the plateau analysis results
    """
    try:
        # Import the data
        df_original, df_processed = load_and_prepare_data_plateau(test_data)

        # Calculate the time length of the plateau
        minutes, seconds = compute_plateau_duration(df_original)

        # Find the most accurate one : the smaller residuals value
        min_form_first, z = compute_mse(df_original)

        # Linear form of the plateau
        if min_form_first == 0 :

            trend_plot, title_trend, shape = create_linear_trend(df_original, z[0][0])
            title_trend += f" - Durée : {str(int(minutes))}min {str(int(seconds))}s"

        # Parabolic form of the plateau
        elif min_form_first == 1 :

            # Parabolic plot of the modelization
            trend_plot = create_parabolic_trend(df_original, z[1][0])
            title_trend = f'Forme parabolique - Durée : {str(int(minutes))}min {str(int(seconds))}s'
            shape = 'parabolique'

        # Scatter plot of the VO2 data
        original_scatter_plot = go.Scatter(
            x=df_original['t'], 
            y=df_original['VO2'], 
            mode="markers",
            name='Données de VO2 brutes', 
            marker=dict(color="purple", size=5)
        )

        # Modelization of the VO2 data along the time
        pwlf_results_processed = perform_pwlf_fit(df_processed["timestamp"], df_processed['VO2'], segments=3)
        y_vo2_processed = compute_y_predicted(
            pwlf_results_processed['slopes'], 
            pwlf_results_processed['breaks'], 
            pwlf_results_processed['intercepts']
        )

        # Get the intersection of the two thresholds
        s1 = intersect(
            round(float(vo2_results[0])/1000, 3), 
            pwlf_results_processed['slopes'], 
            pwlf_results_processed['breaks'], 
            pwlf_results_processed['intercepts'], 
            y_vo2_processed
        )
        s2 = intersect(
            round(float(vo2_results[1])/1000, 3), 
            pwlf_results_processed['slopes'],
            pwlf_results_processed['breaks'], 
            pwlf_results_processed['intercepts'], 
            y_vo2_processed
        )
        x_axis_intersect = []
        x_axis_intersect.append(df_processed['t'].iloc[df_processed['timestamp'].searchsorted(s1[0])])
        x_axis_intersect.append(df_processed['t'].iloc[df_processed['timestamp'].searchsorted(s2[0])])

        # Create the plot for the intersection plot
        processed_scatter_plot = go.Scatter(
            x=df_processed['t'], 
            y=df_processed['VO2'], 
            mode="markers",
            name = 'Données de VO2 traitées', 
            marker=dict(color="darkturquoise", size=5)
        )
        model_plot = go.Scatter(
            x=[df_processed['t'].iloc[df_processed['timestamp'].searchsorted(b)] for b in pwlf_results_processed['breaks']], 
            y=y_vo2_processed, 
            mode="lines",
            name='Modélisation linéaire par morceaux', 
            line=dict(color="royalblue", width=2)
        )
        intersection_plot = go.Scatter(
            x=x_axis_intersect, 
            y=[s1[1], s2[1]], 
            mode="markers",
            name="Points d'intersection", 
            marker=dict(color="black", size=10)
        )

        # Cut the dataframes and create the time axis
        index_s1 = df_processed['timestamp'].searchsorted(s1[0])
        index_plateau = df_processed['timestamp'].searchsorted(df_original['timestamp'].iloc[0])
        df_inter = df_processed[df_processed.index > index_s1]
        df_inter = df_inter[df_inter.index < index_plateau].reset_index(drop = True)
        df_inter["timestamp"] = pd.to_datetime(df_inter['t'], format = "mixed").astype('int64') // 10**9
        df_plateau = df_processed[df_processed.index > index_plateau].reset_index(drop = True)
        df_plateau["timestamp"] = pd.to_datetime(df_plateau['t'], format = "mixed").astype('int64') // 10**9

        # Find the most accurate one : the smaller residuals value
        min_form_second, z = compute_mse(df_plateau)

        # Scatter the VO2
        inter = np.polyfit(df_inter['timestamp'], df_inter['VO2'], 1, full = True)
        threshold_scatter_plot = go.Scatter(
            x=df_inter['t'], 
            y=df_inter['VO2'], 
            mode="markers",
            name = 'Données de VO2 traitées', 
            marker=dict(color="darkturquoise", size=5),
            showlegend=False
        )

        # Plot the linear trend between thresholds
        xbis_1 = np.linspace(df_inter['timestamp'].min(),min(df_plateau['timestamp']),len(df_inter['timestamp'])+1)
        ybis_1 = inter[0][1] + inter[0][0]*xbis_1
        xbis_axis_1 = pd.date_range(df_inter['t'].iloc[0], df_plateau['t'].iloc[0], len(df_inter['t'])+1)
        xbis_2 = np.linspace(min(df_plateau['timestamp']),max(df_plateau['timestamp']),len(df_plateau['timestamp']))
        ybis_2 = inter[0][1] + inter[0][0]*xbis_2
        xbis_axis_2 = pd.date_range(df_plateau['t'].iloc[0], df_plateau['t'].iloc[-1], len(df_plateau['t']))

        trend_threshold_plot = go.Scatter(
            x=xbis_axis_1,
            y=ybis_1, 
            mode="lines",
            name='Tendance linéaire entre les seuils',
            line=dict(color="red", width=2)
        )

        trend_threshold_plateau_plot = go.Scatter(
            x=xbis_axis_2,
            y=ybis_2, 
            mode="lines",
            showlegend=False,
            name='Tendance linéaire entre les seuils',
            line=dict(color="red", width=2)
        )

        # Linear form of the plateau
        if min_form_second == 0 :

            # Linear plot and area values of the trend after the second threshold
            x = df_plateau["timestamp"]
            y = z[0][0][1] + z[0][0][0]*x
            yter = inter[0][1] + inter[0][0]*x
            diff12 = lambda x :z[0][0][1] + z[0][0][0]*x - inter[0][1] - inter[0][0]*x
            diff21 = lambda x :inter[0][1] + inter[0][0]*x - z[0][0][1] - z[0][0][0]*x
            trend_plateau_plot = go.Scatter(
                x=df_plateau['t'],
                y=y,
                name='Tendance linéaire du plateau',
                fill='tonexty',
                mode="lines",
                line=dict(color="gold", width=2)
            )

        # Parabolic form of the plateau
        elif min_form_second == 1 :

            # Parabolic plot and area values of the trend after the second threshold
            x = df_plateau["timestamp"]
            y = z[1][0][2] + z[1][0][1]*x + z[1][0][0]*x**2
            yter = inter[0][1] + inter[0][0]*x
            diff12 = lambda x :z[1][0][2] + z[1][0][1]*x + z[1][0][0]*x**2 - inter[0][1] - inter[0][0]*x
            diff21 = lambda x :inter[0][1] + inter[0][0]*x - z[1][0][2] - z[1][0][1]*x - z[1][0][0]*x**2
            trend_plateau_plot = go.Scatter(
                x=df_plateau['t'],
                y=y,
                name='Tendance parabolique du plateau',
                fill='tonexty',
                mode="lines",
                line=dict(color="gold", width=2)
            )

        # Scatter plot of the VO2 data
        plateau_scatter_plot = go.Scatter(
            x=df_plateau['t'], 
            y=df_plateau['VO2'], 
            mode="markers",
            name='Données de VO2 traitées', 
            marker=dict(color="darkturquoise", size=5),
            showlegend=False
        )

        # Get the integral values
        res = integral(x, diff12, diff21, y, yter)

        # Compute the time results
        time_under = int(sum((t[1] - t[0] for t in res[0])))
        if len(res[0]) != 0 :
            minutes_under, seconds_under = divmod(time_under, 60)
        else :
            minutes_under, seconds_under = 0, 0
        time_above = int(sum((t[1] - t[0] for t in res[1])))
        if len(res[1]) != 0 :
            minutes_above, seconds_above = divmod(time_above, 60)
        else :
            minutes_above, seconds_above = 0, 0
        first_derivative = round(60*np.mean(np.gradient(y)[:10]), 1)

        # Compute the values
        area_under = int(sum((res[2][i][0] for i in range(len(res[2])))))
        area_above = int(sum((res[3][i][0] for i in range(len(res[3])))))
        total_area = int(area_above - area_under)

        # Création du tableau des résultats
        table_trace = go.Table(
            header=dict(
                values=[],
                fill_color='white',
                line_color='white'
            ),
            cells=dict(
                values=[
                    ['Durée sous la tendance linéaire', 
                    'Durée au-dessus de la tendance linéaire',
                    'Aire sous la tendance linéaire', 
                    'Aire au-dessus de la tendance linéaire',
                    'Aire totale', 
                    'Vitesse de début de plateau (~15s)'],
                    [f"{round(minutes_under)}\'{round(seconds_under)}\'\'",
                    f"{round(minutes_above)}\'{round(seconds_above)}\'\'",
                    f"{area_under} s.L/min",
                    f"{area_above} s.L/min",
                    f"{total_area} s.L/min",
                    f"{first_derivative} L/min²"]
                ],
                fill_color=[["#EDEDED", "white"] * 6],  # Alternance des couleurs
                align='center',
                font=dict(size=12),
                height=70
            )
        )

        # Plot the plateau modelizations
        fig = make_subplots(
            rows=2, cols=2,
            vertical_spacing=0.12,
            horizontal_spacing=0.06,
            specs=[[{}, {}], [{}, {"type": "table"}]],
            subplot_titles=(
                title_trend,
                "Points d'intersection\nentre seuils et modélisation",
                'Aire entre la tendance linéaire\ninter-seuils et la VO2',
                None
            )
        )

        # Register the different plots
        plots = [
            [original_scatter_plot, trend_plot], 
            [processed_scatter_plot, model_plot, intersection_plot],
            [threshold_scatter_plot, trend_threshold_plot, trend_threshold_plateau_plot, trend_plateau_plot, plateau_scatter_plot],
            [table_trace]
        ]
        for i, plot in enumerate(plots):
            r = 1 if i <= 1 else 2
            c = 2 if i in [1, 3] else 1
            for p in plot:
                fig.add_trace(p, row=r, col=c)
                if r == 1 and c == 2:
                    fig.add_hline(y=s1[1],row=r,col=c,line=dict(color="brown", width=2, dash="dash"),name='Seuil 1')
                    fig.add_hline(y=s2[1],row=r,col=c,line=dict(color="red", width=2, dash="dash"),name='Seuil 2')

        # Update the axis
        for row in range(1, 3):
            for col in range(1, 3):
                if row == 2 and col == 2:
                    break
                elif row == 1 and col == 1:
                    fig.update_xaxes(
                        title_text="Temps (HH:MM:SS)", 
                        row=row, col=col, showgrid=True)
                else:
                    fig.update_xaxes(
                        title_text="Temps (HH:MM:SS)", 
                        row=row, col=col, showgrid=True)
                fig.update_yaxes(title_text='VO2 (L/min)', row=row, col=col, showgrid=True)

        # Update the layout
        fig.update_layout(
            legend=dict(
                font=dict(size=10),
                yanchor="top",
                y=1.1,
                xanchor="left",
                x=0,
                bgcolor="rgba(255,255,255,0.8)",
                bordercolor="LightGray",
                borderwidth=1,
                orientation="h"
            ),
            template="plotly_white",
            margin=dict(r=30, t=40, l=70, b=60),
            height=1000,
            width=1400
        )

        # Save the plot and the results into the database
        results_plateau = {
            'time_under_trend_plateau': time_under,
            'time_above_trend_plateau': time_above,
            'total_time': int(time_under + time_above),
            'area_under_trend_plateau': area_under,
            'area_above_trend_plateau': area_above,
            'total_area_plateau': total_area,
            'start_speed_plateau': first_derivative,
            'shape_plateau': shape
        }
        
        return fig, results_plateau
    
    except Exception as e:
        logging.error(f"Error in compute_plateau: {e}\n{traceback.format_exc()}")
        raise

def load_and_prepare_data_plateau(test_data):
    """ Load and prepare dataframes with timestamps.

    Parameters
    ----------
    test_data : Test
        Test object that contains plateau and computed dataframes as JSON strings.

    Returns
    -------
    df1 : pandas.DataFrame
        DataFrame containing plateau data with timestamp.
    df2 : pandas.DataFrame
        DataFrame containing computed data with timestamp.
    """
    try:
        # Load JSON data into DataFrames
        df1 = pd.read_json(StringIO(test_data.plateau_dataframe))
        df2 = pd.read_json(StringIO(test_data.computed_dataframe))

        # Ensure timestamp column is added correctly
        for df in [df1, df2]:
            df["timestamp"] = pd.to_datetime(df['t'], format="mixed").astype('int64') // 10**9

        return df1, df2
    
    except Exception as e:
        logging.error(f"Error in load_and_prepare_data_plateau: {e}\n{traceback.format_exc()}")
        raise

def compute_plateau_duration(df):
    """ Calculate plateau duration in minutes and seconds.

    Parameters
    ----------
    df : pandas.DataFrame
        DataFrame containing a 'timestamp' column with Unix timestamp values.
    
    Returns
    -------
    tuple
        Duration in minutes and seconds as a tuple (minutes, seconds).
    """
    try:
        # Calculate the difference between the first and last timestamps
        delta = df['timestamp'].iloc[-1] - df['timestamp'].iloc[0]

        return divmod(delta, 60)
    
    except Exception as e:
        logging.error(f"Error in compute_plateau_duration: {e}\n{traceback.format_exc()}")
        raise

def create_linear_trend(df, model):
    """ Generate a linear trend scatter plot.

    Parameters
    ----------
    df : pandas.DataFrame
        DataFrame containing the 'timestamp' and 't' columns.
    model : tuple
        Tuple containing the slope (model[0]) and intercept (model[1]) of the linear model.
    
    Returns
    -------
    plotly.graph_objects.Scatter
        Plotly scatter plot of the linear trend.
    str
        Title of the trend plot.
    str
        Description of the trend shape.
    """
    try:
        # Linear plot of the modelization
        x = [df['timestamp'].iloc[0], df['timestamp'].iloc[-1]]
        y = model[1] + model[0] * np.array(x)
        trend_plot = go.Scatter(x=[df['t'].iloc[0], df['t'].iloc[-1]], y=y, mode="lines", line=dict(color="limegreen", width=2), name='Tendance linéaire')

        # Determine the precise shape of the modelization
        if model[0] < 0.0001 and model[0] >= 0 :
            title_trend = 'Forme constante'
            shape = 'constant'
        elif model[0] >= 0.0001 :
            title_trend = f'Forme linéaire croissante ({str(round(model[0], 3))})'
            shape = 'linéaire croissant'
        else :
            title_trend = f'Forme linéaire décroissante ({str(round(model[0], 3))})'
            shape = 'linéaire décroissant'

        return trend_plot, title_trend, shape
    
    except Exception as e:
        logging.error(f"Error in create_linear_trend: {e}\n{traceback.format_exc()}")
        raise

def create_parabolic_trend(df, model):
    """ Generate a parabolic trend scatter plot.

    Parameters
    ----------
    df : pandas.DataFrame
        DataFrame containing the 'timestamp' and 't' columns.
    model : tuple
        Tuple containing the coefficients [a, b, c] for the quadratic model (y = ax^2 + bx + c).
    
    Returns
    -------
    plotly.graph_objects.Scatter
        Plotly scatter plot of the parabolic trend.
    """
    try:
        # Generate the parabolic trend plot
        x = np.linspace(df['timestamp'].min(), df['timestamp'].max(), 100)
        y = model[2] + model[1] * x + model[0] * x**2
        x_axis = pd.date_range(df['t'].iloc[0], df['t'].iloc[-1], 100)

        trend_plot = go.Scatter(x=x_axis, y=y, mode="lines", line=dict(color="limegreen", width=2), name='Tendance parabolique')

        return trend_plot
    
    except Exception as e:
        logging.error(f"Error in create_parabolic_trend: {e}\n{traceback.format_exc()}")
        raise
